"""Сверка преподавателей из Excel со списком пользователей prod API.

.. deprecated::
    Используйте ``python data/sync_project_teachers.py`` — единый пайплайн
    парсинга расписания и сверки с prod.
"""

from __future__ import annotations

import json
from pathlib import Path
import sys
import warnings

DATA_DIR = Path(__file__).resolve().parent
sys.path.insert(0, str(DATA_DIR))

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from teacher_matching import build_user_indexes, find_user

USERS_JSON = DATA_DIR / "prod_users.json"
SOURCE_XLSX = DATA_DIR / "project_teachers.xlsx"
OUTPUT_XLSX = DATA_DIR / "project_teachers_marked.xlsx"
SHEET_NAME = "Проектная деятельность"

COL_TEACHER_FIO = 9  # I
COL_IN_SYSTEM = 14  # N
COL_USER_ID = 15  # O
COL_USER_EMAIL = 16  # P


def main() -> None:
    """Отмечает преподавателей из Excel, которые есть в prod."""
    warnings.warn(
        "mark_teachers_in_system.py устарел. "
        "Используйте: python data/sync_project_teachers.py",
        DeprecationWarning,
        stacklevel=2,
    )
    users = json.loads(USERS_JSON.read_text(encoding="utf-8"))
    by_name, by_tokens = build_user_indexes(users)

    workbook = load_workbook(SOURCE_XLSX)
    sheet = workbook[SHEET_NAME]

    header_font = Font(bold=True)
    header_fill = PatternFill(fill_type="solid", fgColor="FFE2EFDA")
    yes_fill = PatternFill(fill_type="solid", fgColor="FFC6EFCE")
    no_fill = PatternFill(fill_type="solid", fgColor="FFFFC7CE")

    sheet.cell(row=1, column=COL_IN_SYSTEM, value="В системе PD")
    sheet.cell(row=1, column=COL_USER_ID, value="ID в PD")
    sheet.cell(row=1, column=COL_USER_EMAIL, value="Email в PD")
    for col in (COL_IN_SYSTEM, COL_USER_ID, COL_USER_EMAIL):
        cell = sheet.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill

    found = 0
    missing = 0
    empty = 0

    for row in range(2, sheet.max_row + 1):
        teacher_name = sheet.cell(row=row, column=COL_TEACHER_FIO).value
        if not teacher_name or not str(teacher_name).strip():
            empty += 1
            continue

        user = find_user(str(teacher_name), by_name=by_name, by_tokens=by_tokens)
        in_system_cell = sheet.cell(row=row, column=COL_IN_SYSTEM)
        id_cell = sheet.cell(row=row, column=COL_USER_ID)
        email_cell = sheet.cell(row=row, column=COL_USER_EMAIL)

        if user:
            found += 1
            in_system_cell.value = "да"
            id_cell.value = user["id"]
            email_cell.value = user.get("email") or ""
            in_system_cell.fill = yes_fill
        else:
            missing += 1
            in_system_cell.value = "нет"
            id_cell.value = ""
            email_cell.value = ""
            in_system_cell.fill = no_fill

    workbook.save(OUTPUT_XLSX)
    print(f"Пользователей в API: {len(users)}")
    print(f"Строк с преподавателем: {found + missing}")
    print(f"Найдено в системе: {found}")
    print(f"Не найдено: {missing}")
    print(f"Пустых строк (без ФИО): {empty}")
    print(f"Сохранено: {OUTPUT_XLSX}")


if __name__ == "__main__":
    main()
