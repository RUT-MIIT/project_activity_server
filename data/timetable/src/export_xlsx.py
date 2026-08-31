from __future__ import annotations

from collections.abc import Callable
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from src.models import GroupResult

HEADERS = [
    "Институт",
    "Аббр. института",
    "Курс",
    "Специальность",
    "Аббр. специальности",
    "Группа",
    "ID группы",
    "Семестр",
    "Преподаватель (ФИО)",
    "Преподаватель (кратко)",
    "ID преподавателя",
    "Кол-во пар",
    "Статус",
]

MARKED_HEADERS = HEADERS + ["В системе PD", "ID в PD", "Email в PD"]

COLUMN_WIDTHS = [35, 12, 8, 45, 12, 14, 12, 22, 35, 18, 14, 10, 18]
MARKED_COLUMN_WIDTHS = COLUMN_WIDTHS + [14, 10, 30]

HEADER_FILL = PatternFill(fill_type="solid", fgColor="FFE2EFDA")
YES_FILL = PatternFill(fill_type="solid", fgColor="FFC6EFCE")
NO_FILL = PatternFill(fill_type="solid", fgColor="FFFFC7CE")


def _group_columns(result: GroupResult) -> list:
    group = result.group
    return [
        group.institute,
        group.institute_abbr,
        group.course,
        group.specialty,
        group.specialty_abbr,
        group.group_name,
        group.group_id,
        result.semester,
    ]


def export_to_xlsx(results: list[GroupResult], output_path: str) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Проектная деятельность"

    sheet.append(HEADERS)
    for cell in sheet[1]:
        cell.font = Font(bold=True)

    for result in results:
        if result.teachers:
            for teacher in result.teachers:
                sheet.append(
                    _group_columns(result)
                    + [
                        teacher.full_fio,
                        teacher.short_fio,
                        teacher.id,
                        teacher.lesson_count,
                        result.status,
                    ]
                )
        else:
            sheet.append(
                _group_columns(result)
                + ["", "", "", result.lesson_count, result.status]
            )

    for index, width in enumerate(COLUMN_WIDTHS, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width

    workbook.save(output_path)


def export_marked_xlsx(
    results: list[GroupResult],
    output_path: str,
    match_teacher: Callable[[str | None], dict[str, Any] | None],
) -> tuple[int, int]:
    """Экспортирует результаты парсинга с колонками сверки с PD."""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Проектная деятельность"

    sheet.append(MARKED_HEADERS)
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL

    found = 0
    missing = 0

    for result in results:
        if result.teachers:
            for teacher in result.teachers:
                user = match_teacher(teacher.full_fio)
                row = _group_columns(result) + [
                    teacher.full_fio,
                    teacher.short_fio,
                    teacher.id,
                    teacher.lesson_count,
                    result.status,
                ]
                if user:
                    found += 1
                    row.extend(["да", user["id"], user.get("email") or ""])
                    in_system_fill = YES_FILL
                else:
                    missing += 1
                    row.extend(["нет", "", ""])
                    in_system_fill = NO_FILL
                sheet.append(row)
                in_system_cell = sheet.cell(row=sheet.max_row, column=len(HEADERS) + 1)
                in_system_cell.fill = in_system_fill
        else:
            sheet.append(
                _group_columns(result)
                + ["", "", "", result.lesson_count, result.status, "", "", ""]
            )

    for index, width in enumerate(MARKED_COLUMN_WIDTHS, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width

    workbook.save(output_path)
    return found, missing
