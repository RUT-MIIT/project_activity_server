from __future__ import annotations

import argparse
from pathlib import Path
import re
from typing import Any

from openpyxl import load_workbook
import pandas as pd


def _normalize_header(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip().lower()


def _extract_group_abbrev_from_text(text: str) -> str | None:
    """
    Пытается извлечь аббревиатуру группы из строки вида "Группа : XXX(....)".
    Возвращает аббревиатуру или None.
    """
    cleaned = text.strip()
    if not cleaned:
        return None

    if "группа" not in cleaned.lower() or ":" not in cleaned:
        return None

    match = re.search(r":\s*([^\(\r\n]+)", cleaned)
    if not match:
        return None

    candidate = match.group(1).strip()
    if not candidate:
        return None

    abbrev = candidate.split()[0].strip()
    return abbrev or None


def _looks_like_student_id(value: Any) -> bool:
    if value is None:
        return False
    if isinstance(value, (int, float)) and not pd.isna(value):
        return True
    if isinstance(value, str):
        v = value.strip()
        return v.isdigit()
    return False


def extract_abbrev_column(input_path: Path, output_path: Path) -> None:
    """
    Извлекает столбец "Аббревеатура группа" из Excel-файла и сохраняет новый файл,
    убирая пустые значения.
    """
    # 1) Пытаемся найти нормальный столбец по заголовку (классический табличный формат).
    workbook = pd.read_excel(input_path, sheet_name=None, dtype=object)
    target_norm = _normalize_header("Аббревеатура группа")
    output_sheets: dict[str, pd.DataFrame] = {}

    for sheet_name, df in workbook.items():
        if df is None or df.empty:
            continue

        normalized_to_original: dict[str, str] = {
            _normalize_header(col): str(col) for col in df.columns
        }
        if target_norm not in normalized_to_original:
            continue

        original_col = normalized_to_original[target_norm]
        out_df = df[[original_col]].copy()
        out_df = out_df.rename(columns={original_col: "Аббревеатура группа"})
        out_df["Аббревеатура группа"] = (
            out_df["Аббревеатура группа"].astype("string").str.strip()
        )
        out_df = out_df.dropna(subset=["Аббревеатура группа"])
        out_df = out_df[out_df["Аббревеатура группа"] != ""]

        if not out_df.empty:
            output_sheets[sheet_name] = out_df

    # 2) Если табличного столбца нет — разбираем "отчётный" формат (в одной ячейке
    #    может быть "Группа : XXX(курс ...)" и т.п.). Собираем список аббревиатур.
    if not output_sheets:
        output_sheets = {}
        raw_workbook = pd.read_excel(
            input_path, sheet_name=None, header=None, dtype=object, engine="openpyxl"
        )

        for sheet_name, df in raw_workbook.items():
            if df is None or df.empty:
                continue

            series = df.stack(future_stack=True).astype("string").str.strip().dropna()
            if series.empty:
                continue

            # Выдёргиваем подстроку после ":" до "(" или конца строки.
            extracted = (
                series.str.extract(r":\s*([^\(\r\n]+)", expand=False)
                .astype("string")
                .str.strip()
            )
            extracted = extracted.dropna()
            extracted = extracted[extracted != ""]

            # Часто после ":" идёт именно аббревиатура группы, а дальше в скобках курс.
            # Оставляем только "первый токен" до пробела на случай хвостов.
            extracted = extracted.str.split().str[0].astype("string").str.strip()
            extracted = extracted.dropna()
            extracted = extracted[extracted != ""]

            extracted = extracted.drop_duplicates().reset_index(drop=True)
            if not extracted.empty:
                output_sheets[sheet_name] = pd.DataFrame(
                    {"Аббревеатура группа": extracted}
                )

    if not output_sheets:
        raise ValueError(
            'Не удалось извлечь "Аббревеатура группа": ни столбец, ни строки вида '
            '"Группа : ...(...)" не найдены.'
        )

    output_path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        for sheet_name, out_df in output_sheets.items():
            out_df.to_excel(writer, sheet_name=sheet_name, index=False)


def add_abbrev_column_to_students(input_path: Path, output_path: Path) -> None:
    """
    Создаёт копию Excel со студентами, добавляя столбец "Аббревеатура группа"
    и заполняя его для строк студентов.
    """
    wb = load_workbook(filename=input_path)

    for ws in wb.worksheets:
        max_col = ws.max_column
        max_row = ws.max_row

        header_row_idx: int | None = None
        id_col_idx: int | None = None
        current_group_abbrev: str | None = None

        # Находим строку заголовков (где встречается ID_student) и колонку ID.
        for r in range(1, max_row + 1):
            for c in range(1, max_col + 1):
                v = ws.cell(row=r, column=c).value
                if isinstance(v, str) and v.strip() == "ID_student":
                    header_row_idx = r
                    id_col_idx = c
                    break
            if header_row_idx is not None:
                break

        # Если это не “лист со студентами”, пропускаем.
        if header_row_idx is None or id_col_idx is None:
            continue

        new_col_idx = max_col + 1
        ws.cell(row=header_row_idx, column=new_col_idx).value = "Аббревеатура группа"

        # Идём по строкам ниже заголовка, отслеживаем текущую группу и проставляем в строках студентов.
        for r in range(header_row_idx + 1, max_row + 1):
            # Обновляем текущую группу, если в строке есть маркер "Группа : ..."
            for c in range(1, max_col + 1):
                v = ws.cell(row=r, column=c).value
                if isinstance(v, str) and v.strip():
                    maybe = _extract_group_abbrev_from_text(v)
                    if maybe:
                        current_group_abbrev = maybe
                        break

            # Если это строка студента (есть ID_student), проставляем группу.
            student_id = ws.cell(row=r, column=id_col_idx).value
            if _looks_like_student_id(student_id):
                ws.cell(row=r, column=new_col_idx).value = current_group_abbrev

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description=(
            "Работа с Excel групп/студентов: либо выделить аббревиатуры, "
            "либо добавить столбец с аббревиатурой в копию файла."
        )
    )
    parser.add_argument(
        "--input",
        type=Path,
        default=Path("groups_01_09.xlsx"),
        help="Путь к исходному Excel-файлу.",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=Path("groups_01_09_with_abbrev.xlsx"),
        help="Путь к выходному Excel-файлу.",
    )
    parser.add_argument(
        "--mode",
        choices=["extract", "add-column"],
        default="add-column",
        help=(
            'Режим: "extract" — сохранить только аббревиатуры; '
            '"add-column" — копия со студентами + новый столбец.'
        ),
    )
    return parser


def main() -> None:
    args = build_parser().parse_args()
    if args.mode == "extract":
        extract_abbrev_column(input_path=args.input, output_path=args.output)
        return

    add_abbrev_column_to_students(input_path=args.input, output_path=args.output)


if __name__ == "__main__":
    main()
